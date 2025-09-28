import os
import re
import pandas as pd
from collections import defaultdict
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.prompt import Prompt, IntPrompt
from rich.table import Table
from rich.panel import Panel

console = Console()

# ---- Alias Dictionary (expandable) ---- #
COLUMN_ALIASES = {
    "country": ["ctry", "nation"],
    "product": ["sku", "item", "pdct"],
    "description": ["desc", "details"],
    "code": ["id", "identifier"],
    "amount": ["amt", "value"],
    "month": ["mnth", "period"],
    "region": ["area", "zone"],
    "name": ["nm", "title"],
}

# ---- Utilities ---- #
def load_file(path):
    ext = os.path.splitext(path)[-1].lower()
    if ext == ".csv":
        return pd.read_csv(path)
    elif ext in [".xlsx", ".xls"]:
        return pd.read_excel(path)
    else:
        raise ValueError("Unsupported file type.")

def normalize(text):
    return re.sub(r"[^a-z0-9]", "", text.lower())

def map_columns_by_alias(df1_cols, df2_cols):
    mapping = {}
    for col1 in df1_cols:
        norm_col1 = normalize(col1)
        tokens1 = norm_col1.split()

        for col2 in df2_cols:
            norm_col2 = normalize(col2)
            if norm_col1 == norm_col2:
                mapping[col1] = col2
                break

            # Word-based alias match
            for word in norm_col1.split():
                for key, aliases in COLUMN_ALIASES.items():
                    if word == key or word in aliases:
                        for word2 in norm_col2.split():
                            if word2 == key or word2 in aliases:
                                mapping[col1] = col2
                                break
    return mapping

def get_rows_with_special_characters(df):
    pattern = re.compile(r'[^a-zA-Z0-9.\- ]')
    return df[df.apply(lambda row: row.astype(str).apply(lambda x: bool(pattern.search(x))).any(), axis=1)]

def export_excel(df, path, sheet="Sheet1"):
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet)
        ws = writer.sheets[sheet]
        thin = Side(border_style="thin", color="000000")
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

# ---- Validation Options ---- #
def perform_null_check(df1, df2):
    nulls1 = df1[df1.isnull().any(axis=1)]
    nulls2 = df2[df2.isnull().any(axis=1)]
    if nulls1.empty and nulls2.empty:
        console.print("[green]✅ No null values in either file.[/green]")
    else:
        export_excel(nulls1, "nulls_file1.xlsx", "File1_Nulls")
        export_excel(nulls2, "nulls_file2.xlsx", "File2_Nulls")
        console.print("[yellow]⚠️ Null value records exported to 'nulls_file1.xlsx' and 'nulls_file2.xlsx'.[/yellow]")

def perform_negative_check(df1, df2):
    numeric1 = df1.select_dtypes(include="number")
    numeric2 = df2.select_dtypes(include="number")
    neg1 = df1[(numeric1 < 0).any(axis=1)]
    neg2 = df2[(numeric2 < 0).any(axis=1)]
    if neg1.empty and neg2.empty:
        console.print("[green]✅ No negative values in either file.[/green]")
    else:
        export_excel(neg1, "negatives_file1.xlsx", "File1_Negatives")
        export_excel(neg2, "negatives_file2.xlsx", "File2_Negatives")
        console.print("[yellow]⚠️ Negative value records exported to 'negatives_file1.xlsx' and 'negatives_file2.xlsx'.[/yellow]")

def perform_special_char_check(df1, df2):
    special1 = get_rows_with_special_characters(df1)
    special2 = get_rows_with_special_characters(df2)
    if special1.empty and special2.empty:
        console.print("[green]✅ No special characters in either file.[/green]")
    else:
        export_excel(special1, "special_chars_file1.xlsx", "File1_SpecialChars")
        export_excel(special2, "special_chars_file2.xlsx", "File2_SpecialChars")
        console.print("[yellow]⚠️ Special character rows exported to 'special_chars_file1.xlsx' and 'special_chars_file2.xlsx'.[/yellow]")

# ---- Comparison ---- #
def row_comparison(df1, df2):
    df1['__row__'] = df1.astype(str).agg('|'.join, axis=1)
    df2['__row__'] = df2.astype(str).agg('|'.join, axis=1)
    missing = df1[~df1['__row__'].isin(df2['__row__'])].drop(columns=['__row__'])
    extra = df2[~df2['__row__'].isin(df1['__row__'])].drop(columns=['__row__'])
    export_excel(missing, "missing_in_file2.xlsx", "MissingInFile2")
    export_excel(extra, "extra_in_file2.xlsx", "ExtraInFile2")
    console.print("[green]Row-by-row comparison completed. Results in Excel files.[/green]")

def cell_by_cell_comparison(df1, df2, mapping, keys):
    df1 = df1.rename(columns=mapping)
    df2 = df2.rename(columns=mapping)

    merged = df1.merge(df2, on=keys, suffixes=('_f1', '_f2'), how='outer', indicator=True)
    diffs = []
    for col in df1.columns:
        if col not in keys and f"{col}_f1" in merged and f"{col}_f2" in merged:
            merged[f"{col}_diff"] = merged[f"{col}_f1"] != merged[f"{col}_f2"]
            diffs.append(f"{col}_diff")

    differences = merged[merged[diffs].any(axis=1)]

    if differences.empty:
        console.print("[green]✅ No cell-level differences found.[/green]")
        return

    report_rows = []
    for _, row in differences.iterrows():
        key_data = {k: row[k] for k in keys}
        for col in df1.columns:
            if col not in keys:
                v1 = row.get(f"{col}_f1", "")
                v2 = row.get(f"{col}_f2", "")
                if v1 != v2:
                    report_rows.append({**key_data, "Column": col, "File1": v1, "File2": v2})

    result_df = pd.DataFrame(report_rows)
    export_excel(result_df, "cell_level_differences.xlsx", "CellDiffs")
    console.print("[yellow]Cell-level differences saved to 'cell_level_differences.xlsx'[/yellow]")

# ---- Main ---- #
def main():
    console.print(Panel("[bold cyan]🧠 Data Validation and Comparison Tool[/bold cyan]"))

    path1 = Prompt.ask("[bold]📂 Enter path to File 1[/bold]")
    path2 = Prompt.ask("[bold]📂 Enter path to File 2[/bold]")
    df1 = load_file(path1)
    df2 = load_file(path2)

    console.print("[bold magenta]\n🔗 Auto Mapping Columns...[/bold magenta]")
    mapping = map_columns_by_alias(df1.columns, df2.columns)

    table = Table(title="Mapped Columns")
    table.add_column("File 1 Column")
    table.add_column("File 2 Column")
    for k, v in mapping.items():
        table.add_row(k, v)
    console.print(table)

    while True:
        console.print("\n[bold cyan]Choose Validation/Comparison Option:[/bold cyan]")
        console.print("1. Count of Records")
        console.print("2. Null Value Records")
        console.print("3. Negative Value Records")
        console.print("4. Special Character Records")
        console.print("5. Row-by-Row Comparison")
        console.print("6. Cell-by-Cell Comparison")
        console.print("7. Exit")

        choice = IntPrompt.ask("Enter your choice", choices=[str(i) for i in range(1, 8)])

        if choice == 1:
            console.print(f"File 1 Rows: {len(df1)}, File 2 Rows: {len(df2)}")
        elif choice == 2:
            perform_null_check(df1, df2)
        elif choice == 3:
            perform_negative_check(df1, df2)
        elif choice == 4:
            perform_special_char_check(df1, df2)
        elif choice == 5:
            df1_m = df1.rename(columns=mapping)
            df2_m = df2.rename(columns=mapping)
            row_comparison(df1_m, df2_m)
        elif choice == 6:
            keys = Prompt.ask("Enter unique key columns (comma-separated)").split(",")
            keys = [k.strip() for k in keys if k.strip()]
            df1_m = df1.rename(columns=mapping)
            df2_m = df2.rename(columns=mapping)
            cell_by_cell_comparison(df1, df2, mapping, keys)
        elif choice == 7:
            console.print("[bold green]✅ Exiting. All done![/bold green]")
            break

if __name__ == "__main__":
    main()
