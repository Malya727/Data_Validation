import pandas as pd
import openpyxl
from colorama import init, Fore, Style
from openpyxl.styles import Font, Border, Side
import warnings
import os
import time
import re
from collections import defaultdict

init(autoreset=True)
warnings.filterwarnings('ignore')

def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

def apply_borders_to_all_cells(ws):
    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def get_file_size(file_path):
    return os.path.getsize(file_path) / 1024  # Convert bytes to kilobytes

def find_unique_identifier(df):
    for col in df.columns:
        if df[col].is_unique:
            return col
    return df.index.name if df.index.name else 'index'

def read_file(file_path):
    try:
        if file_path.endswith('.csv'):
            return pd.read_csv(file_path)
        elif file_path.endswith('.xlsx'):
            return pd.read_excel(file_path)
        else:
            raise ValueError("Unsupported file format. Please provide a .csv or .xlsx file.")
    except FileNotFoundError:
        print(Fore.RED + f"Error: The file '{file_path}' was not found.")
        return None

def identify_columns_with_negative_values(df):
    columns_with_negative_values = []
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]) and (df[col] < 0).any():
            columns_with_negative_values.append(col)
    return columns_with_negative_values

def identify_columns_with_null_values(df):
    return df.columns[df.isnull().any()].tolist()

def identify_columns_with_special_characters(df):
    special_char_pattern = re.compile(r'[^a-zA-Z0-9.-]')
    columns_with_special_characters = defaultdict(set)
    for col in df.columns:
        for value in df[col].astype(str):
            matches = special_char_pattern.findall(value)
            if matches:
                columns_with_special_characters[col].update(matches)
    return columns_with_special_characters

def get_rows_with_special_characters(df):
    special_char_pattern = re.compile(r'[^a-zA-Z0-9.]')
    rows_with_special_characters = df[df.apply(lambda row: row.astype(str).apply(lambda x: bool(special_char_pattern.search(x))).any(), axis=1)]
    return rows_with_special_characters

def convert_month_column_to_datetime(df, column_name):
    try:
        df[column_name] = pd.to_datetime(df[column_name], format='%d-%b', errors='coerce')
    except Exception as e:
        print(Fore.RED + f"Error converting column '{column_name}' to datetime: {e}")

def convert_month_column_to_string(df, column_name):
    try:
        df[column_name] = df[column_name].dt.strftime('%d-%b')
    except Exception as e:
        print(Fore.RED + f"Error converting column '{column_name}' to string: {e}")

def compare_excel_files(source_file, target_file):
    start_time = time.time()
    df_source = read_file(source_file)
    source_load_time = time.time() - start_time
    if df_source is None:
        return

    print(Fore.GREEN + "Source file successfully loaded..")

    start_time = time.time()
    df_target = read_file(target_file)
    target_load_time = time.time() - start_time
    if df_target is None:
        return

    print(Fore.GREEN + "Target file successfully loaded..")

    source_id_col = find_unique_identifier(df_source)
    target_id_col = find_unique_identifier(df_target)

    if source_id_col not in df_source.columns:
        df_source.reset_index(inplace=True)
        source_id_col = 'index'
    if target_id_col not in df_target.columns:
        df_target.reset_index(inplace=True)
        target_id_col = 'index'

    # Convert Month columns to datetime if they exist
    if 'Month' in df_source.columns:
        convert_month_column_to_datetime(df_source, 'Month')
    if 'Month' in df_target.columns:
        convert_month_column_to_datetime(df_target, 'Month')

    source_file_size = get_file_size(source_file)
    target_file_size = get_file_size(target_file)
    source_columns = df_source.columns.tolist()
    target_columns = df_target.columns.tolist()
    source_column_count = len(source_columns)
    target_column_count = len(target_columns)

    negative_value_columns_source = identify_columns_with_negative_values(df_source)
    negative_value_columns_target = identify_columns_with_negative_values(df_target)

    null_value_columns_source = identify_columns_with_null_values(df_source)
    null_value_columns_target = identify_columns_with_null_values(df_target)

    special_char_columns_source = identify_columns_with_special_characters(df_source)
    special_char_columns_target = identify_columns_with_special_characters(df_target)

    print(Fore.GREEN + "\nSummary of Source File:")
    print(Fore.YELLOW + f"File Size: {source_file_size:.2f} KB")
    print(Fore.YELLOW + f"Loading Time: {source_load_time:.2f} seconds")
    print(Fore.YELLOW + f"Column Names: {source_columns}")
    print(Fore.YELLOW + f"Column Count: {source_column_count}")

    if negative_value_columns_source:
        print(Fore.YELLOW + f"Columns with negative values: {negative_value_columns_source}")
    if null_value_columns_source:
        print(Fore.YELLOW + f"Columns with null values: {null_value_columns_source}")
    if special_char_columns_source:
        for col, chars in special_char_columns_source.items():
            print(Fore.YELLOW + f"Column '{col}' contains special characters: {', '.join(chars)}")

    print(Fore.GREEN + "\nSummary of Target File:")
    print(Fore.YELLOW + f"File Size: {target_file_size:.2f} KB")
    print(Fore.YELLOW + f"Loading Time: {target_load_time:.2f} seconds")
    print(Fore.YELLOW + f"Column Names: {target_columns}")
    print(Fore.YELLOW + f"Column Count: {target_column_count}")

    if negative_value_columns_target:
        print(Fore.YELLOW + f"Columns with negative values: {negative_value_columns_target}")
    if null_value_columns_target:
        print(Fore.YELLOW + f"Columns with null values: {null_value_columns_target}")
    if special_char_columns_target:
        for col, chars in special_char_columns_target.items():
            print(Fore.YELLOW + f"Column '{col}' contains special characters: {', '.join(chars)}")

    df_source.set_index(source_id_col, inplace=True)
    df_target.set_index(target_id_col, inplace=True)
    print(Fore.YELLOW + "\nWhat type of reconciliation you want to perform?\n1.Count of records\n2.Data Comparison\n3.Null value Records\n4.Negative value Records\n5.Special characters Records\nChoose reconciliation type? (Enter number:) ", end='')
    y = input(Fore.WHITE)

    if y == "1":
        source_record_count = len(df_source)
        print(Fore.YELLOW + f"Count of records in the Source file: {source_record_count}")
        target_record_count = len(df_target)
        print(Fore.YELLOW + f"Count of records in the Target file: {target_record_count}")

        extra_source_rows = df_source[~df_source.index.isin(df_target.index)]
        extra_target_rows = df_target[~df_target.index.isin(df_source.index)]

        if source_record_count == target_record_count and extra_source_rows.empty and extra_target_rows.empty:
            print(Fore.GREEN + "All records match. No discrepancies found.")
        else:
            with pd.ExcelWriter('Mismatch_Records.xlsx', engine='openpyxl') as writer:
                if not extra_source_rows.empty:
                    if 'Month' in extra_source_rows.columns:
                        convert_month_column_to_string(extra_source_rows, 'Month')
                    extra_source_rows.reset_index().to_excel(writer, sheet_name='Extra_Source_Rows', index=False)
                    print(Fore.RED + "Extra rows in Source file have been saved to 'Extra_Source_Rows' sheet.")

                if not extra_target_rows.empty:
                    if 'Month' in extra_target_rows.columns:
                        convert_month_column_to_string(extra_target_rows, 'Month')
                    extra_target_rows.reset_index().to_excel(writer, sheet_name='Extra_Target_Rows', index=False)
                    print(Fore.RED + "Extra rows in Target file have been saved to 'Extra_Target_Rows' sheet.")

                workbook = writer.book
                for sheetname in workbook.sheetnames:
                    ws = workbook[sheetname]
                    auto_adjust_column_width(ws)
                    apply_borders_to_all_cells(ws)

            print(Fore.GREEN + "All results have been saved to 'Mismatch_Records.xlsx'.")

    elif y == "2":
        common_index = df_source.index.intersection(df_target.index)
        df_source_common = df_source.loc[common_index]
        df_target_common = df_target.loc[common_index]

        matched_columns = {}
        unmatched_source_cols = []
        print(Fore.YELLOW + "Comparing source and target files.")
        for source_col in df_source_common.columns:
            match_counts = df_target_common.apply(lambda col: (col.values == df_source_common[source_col].values).sum())
            best_match = match_counts.idxmax()

            if match_counts[best_match] > 0:
                matched_columns[source_col] = best_match
            else:
                unmatched_source_cols.append(source_col)

        aligned_target = df_target_common[matched_columns.values()].copy()
        aligned_target.columns = matched_columns.keys()

        comparison = df_source_common[matched_columns.keys()].compare(aligned_target)
        mismatched_rows = comparison[comparison.apply(lambda row: any(row), axis=1)]

        # Convert datetime columns back to original format for output
        if 'Month' in mismatched_rows.columns.get_level_values(0):
            mismatched_rows[('Month', 'self')] = mismatched_rows[('Month', 'self')].dt.strftime('%d-%b')
            mismatched_rows[('Month', 'other')] = mismatched_rows[('Month', 'other')].dt.strftime('%d-%b')

        if mismatched_rows.empty:
            print(Fore.GREEN + "Data Matched. No discrepancies found.")
        else:
            with pd.ExcelWriter('Mismatch_Differences.xlsx', engine='openpyxl') as writer:
                print(Fore.RED + "Differences found:")
                mismatched_rows.reset_index(inplace=True)
                mismatched_rows.columns = ['_'.join(col).strip() if isinstance(col, tuple) else col for col in mismatched_rows.columns.values]
                mismatched_rows.to_excel(writer, sheet_name='Mismatch_Differences', index=False)
                print(Fore.RED + "Mismatched data with identifiers has been saved to 'Mismatch_Differences' sheet.")

                if unmatched_source_cols:
                    print(Fore.YELLOW + "Extra columns in source file:")
                    print(unmatched_source_cols)

                extra_target_cols = [col for col in df_target.columns if col not in matched_columns.values()]
                if extra_target_cols:
                    print(Fore.YELLOW + "Extra columns in target file:")
                    print(extra_target_cols)

                print(Fore.GREEN + "Processing completed!!!")

                workbook = writer.book
                for sheetname in workbook.sheetnames:
                    ws = workbook[sheetname]
                    auto_adjust_column_width(ws)
                    apply_borders_to_all_cells(ws)

    elif y == "3":
        source_null_values = df_source[df_source.isnull().any(axis=1)]
        target_null_values = df_target[df_target.isnull().any(axis=1)]

        if source_null_values.empty and target_null_values.empty:
            print(Fore.GREEN + "No null values found in both source and target files.")
        else:
            with pd.ExcelWriter('Null_Values_Records.xlsx', engine='openpyxl') as writer:
                if not source_null_values.empty:
                    if 'Month' in source_null_values.columns:
                        convert_month_column_to_string(source_null_values, 'Month')
                    source_null_values.reset_index().to_excel(writer, sheet_name='Source_Null_Values', index=False)
                if not target_null_values.empty:
                    if 'Month' in target_null_values.columns:
                        convert_month_column_to_string(target_null_values, 'Month')
                    target_null_values.reset_index().to_excel(writer, sheet_name='Target_Null_Values', index=False)

                workbook = writer.book
                for sheetname in workbook.sheetnames:
                    ws = workbook[sheetname]
                    auto_adjust_column_width(ws)
                    apply_borders_to_all_cells(ws)

            print(Fore.GREEN + "Rows with null values have been saved to 'Null_Values_Records.xlsx'.")

    elif y == "4":
        source_negative_values = df_source.select_dtypes(include=[int, float]).lt(0).any(axis=1)
        target_negative_values = df_target.select_dtypes(include=[int, float]).lt(0).any(axis=1)

        source_negative_values = df_source[source_negative_values]
        target_negative_values = df_target[target_negative_values]

        if source_negative_values.empty and target_negative_values.empty:
            print(Fore.GREEN + "No negative values found in both source and target files.")
        else:
            with pd.ExcelWriter('Negative_Values_Records.xlsx', engine='openpyxl') as writer:
                if not source_negative_values.empty:
                    if 'Month' in source_negative_values.columns:
                        convert_month_column_to_string(source_negative_values, 'Month')
                    source_negative_values.reset_index().to_excel(writer, sheet_name='Source_Negative_Values', index=False)
                if not target_negative_values.empty:
                    if 'Month' in target_negative_values.columns:
                        convert_month_column_to_string(target_negative_values, 'Month')
                    target_negative_values.reset_index().to_excel(writer, sheet_name='Target_Negative_Values', index=False)

                workbook = writer.book
                for sheetname in workbook.sheetnames:
                    ws = workbook[sheetname]
                    auto_adjust_column_width(ws)
                    apply_borders_to_all_cells(ws)

            print(Fore.GREEN + "Rows with negative values have been saved to 'Negative_Values_Records.xlsx'.")

    elif y == "5":
        source_special_char_rows = get_rows_with_special_characters(df_source)
        print(source_special_char_rows)
        target_special_char_rows = get_rows_with_special_characters(df_target)
       

        if source_special_char_rows.empty and target_special_char_rows.empty:
            print(Fore.GREEN + "No special characters found in both source and target files.")
        else:
            with pd.ExcelWriter('Special_Characters_Records.xlsx', engine='openpyxl') as writer:
                if not source_special_char_rows.empty:
                    if 'Month' in source_special_char_rows.columns:
                        convert_month_column_to_string(source_special_char_rows, 'Month')
                    source_special_char_rows.reset_index().to_excel(writer, sheet_name='Source_Special_Characters', index=False)
                if not target_special_char_rows.empty:
                    if 'Month' in target_special_char_rows.columns:
                        convert_month_column_to_string(target_special_char_rows, 'Month')
                    target_special_char_rows.reset_index().to_excel(writer, sheet_name='Target_Special_Characters', index=False)

                workbook = writer.book
                for sheetname in workbook.sheetnames:
                    ws = workbook[sheetname]
                    auto_adjust_column_width(ws)
                    apply_borders_to_all_cells(ws)

            print(Fore.GREEN + "Rows with special characters have been saved to 'Special_Characters_Records.xlsx'.")

